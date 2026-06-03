import os
import sys
import datetime
from collections import defaultdict
import pandas as pd
import pypdf
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_duration(seconds):
    """Converts a duration in seconds into 'DD HH:MM:SS' string format."""
    days = int(seconds // 86400)
    seconds %= 86400
    hours = int(seconds // 3600)
    seconds %= 3600
    minutes = int(seconds // 60)
    seconds = int(seconds % 60)
    return f"{days:02d} {hours:02d}:{minutes:02d}:{seconds:02d}"

def draw_progress_bar(current, total, bar_length=30):
    """Generates and prints a dynamic text-based progress bar."""
    percent = float(current) / total
    arrow = '-' * int(round(percent * bar_length) - 1) + '>'
    spaces = ' ' * (bar_length - len(arrow))
    sys.stdout.write(f"\r    [{arrow}{spaces}] Page {current}/{total} ({int(percent * 100)}%)")
    sys.stdout.flush()

def extract_batch_metadata(reader):
    """Scans the PDF to find the Start Batch and End Batch timestamps."""
    batch_start = None
    batch_end = None
    
    # Check first few and last few pages to quickly extract metadata
    pages_to_check = list(reader.pages[:5]) + list(reader.pages[-5:])
    for page in pages_to_check:
        text = page.extract_text()
        if not text:
            continue
            
        lines = text.split('\n')
        for line in lines:
            if "Start Batch" in line:
                # Expected format segment: "Start Batch 07 May 2026 18:25:44"
                parts = line.split("Start Batch")[-1].strip().split()
                if len(parts) >= 4:
                    try:
                        date_str = f"{parts[0]} {parts[1]} {parts[2]} {parts[3]}"
                        batch_start = datetime.datetime.strptime(date_str, "%d %B %Y %H:%M:%S")
                    except ValueError:
                        pass
            if "End Batch" in line:
                # Expected format segment: "End Batch 08 May 2026 02:52:20"
                parts = line.split("End Batch")[-1].strip().split()
                if len(parts) >= 4:
                    try:
                        date_str = f"{parts[0]} {parts[1]} {parts[2]} {parts[3]}"
                        batch_end = datetime.datetime.strptime(date_str, "%d %B %Y %H:%M:%S")
                    except ValueError:
                        pass
                        
        if batch_start and batch_end:
            break
            
    return batch_start, batch_end

def process_single_pdf(pdf_path, output_excel_path):
    """Parses a single alarm PDF, matches incomplete states with metadata, and formats Excel."""
    section_rows = {'Filler': [], 'Isolator': []}
    valid_states = {'UNACK_ALM', 'ACK_RTN', 'ACK'}
    
    try:
        reader = pypdf.PdfReader(pdf_path)
    except Exception as e:
        print(f"\n  [ERROR] Could not read PDF file {pdf_path}: {e}")
        return False
        
    total_pages = len(reader.pages)
    if total_pages == 0:
        return False

    # 1. Dynamically find Batch Boundaries
    batch_start, batch_end = extract_batch_metadata(reader)
    
    # 2. Extract Alarm Log Rows
    for page_num, page in enumerate(reader.pages, start=1):
        draw_progress_bar(page_num, total_pages)
        text = page.extract_text()
        if not text:
            continue
            
        current_section = None
        if "Alarms Filler" in text:
            current_section = 'Filler'
        elif "Alarms Isolator" in text:
            current_section = 'Isolator'
            
        if not current_section:
            continue
            
        lines = text.split('\n')
        for line in lines:
            parts = line.strip().split()
            if len(parts) >= 6:
                if '.' in parts[0] and parts[2] in valid_states:
                    try:
                        date_str, time_str, state, code = parts[0], parts[1], parts[2], parts[3]
                        user = parts[-1]
                        desc = " ".join(parts[4:-1])
                        dt = datetime.datetime.strptime(f"{date_str} {time_str}", "%d.%m.%Y %H:%M:%S")
                        
                        section_rows[current_section].append({
                            'datetime': dt, 'state': state, 'code': code, 'desc': desc, 'user': user
                        })
                    except Exception:
                        continue
    print()

    # 3. Process Lifecycles and Write to Sheets
    try:
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            for section, rows in section_rows.items():
                rows.reverse()  # Flip to chronological order
                
                active_alarms = {}
                alarm_summary = defaultdict(lambda: {
                    'quantity': 0, 
                    'total_duration_secs': 0.0,
                    'remarks': set()
                })
                
                # To capture alarms from a previous batch, we track keys that clear without an activation sequence first
                cleared_without_start = set()
                
                for row in rows:
                    key = (row['code'], row['desc'])
                    state = row['state']
                    dt = row['datetime']
                    
                    if state == 'UNACK_ALM':
                        active_alarms[key] = dt
                        alarm_summary[key]['quantity'] += 1
                        
                    elif state == 'ACK_RTN':
                        if key in active_alarms:
                            # Balanced scenario
                            start_time = active_alarms[key]
                            duration = (dt - start_time).total_seconds()
                            alarm_summary[key]['total_duration_secs'] += duration
                            alarm_summary[key]['remarks'].add("Completed within batch")
                            del active_alarms[key]
                        else:
                            # Scenario A: Spilled over from previous batch
                            if batch_start:
                                duration = (dt - batch_start).total_seconds()
                                if duration > 0:
                                    alarm_summary[key]['total_duration_secs'] += duration
                            
                            # Increment quantity because it was an active disruption during this batch window
                            alarm_summary[key]['quantity'] += 1
                            alarm_summary[key]['remarks'].add("Alarm active from previous batch (Duration calculated from Batch Start)")
                
                # Scenario B: Unacknowledged alarms lingering open until the end of the batch run
                for key, start_time in active_alarms.items():
                    if batch_end:
                        duration = (batch_end - start_time).total_seconds()
                        if duration > 0:
                            alarm_summary[key]['total_duration_secs'] += duration
                    alarm_summary[key]['remarks'].add("Alarm active until end of batch (Duration calculated to Batch End)")
                
                # Format to final Dataframe structure
                formatted_table = []
                for (code, desc), stats in alarm_summary.items():
                    # Prioritize highlighting anomalies over standard "Completed within batch" remark
                    rem_set = stats['remarks']
                    if len(rem_set) > 1 and "Completed within batch" in rem_set:
                        rem_set.remove("Completed within batch")
                    remark_str = "; ".join(rem_set) if rem_set else "Completed within batch"
                    
                    formatted_table.append({
                        'Alarms / Warnings': f"{code} {desc}",
                        'Quantity': stats['quantity'],
                        'Alarm on': format_duration(stats['total_duration_secs']),
                        'Remarks': remark_str,
                        '_sort_key': stats['quantity']
                    })
                
                if formatted_table:
                    df = pd.DataFrame(formatted_table).sort_values(by='_sort_key', ascending=False).drop(columns=['_sort_key'])
                else:
                    df = pd.DataFrame(columns=['Alarms / Warnings', 'Quantity', 'Alarm on', 'Remarks'])
                
                sheet_name = f"Alarms {section}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply Style Palette
                ws = writer.sheets[sheet_name]
                ws.views.sheetView[0].showGridLines = True
                
                header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
                zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                body_font = Font(name="Segoe UI", size=10)
                
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )
                
                ws.row_dimensions[1].height = 26
                for col_idx in range(1, 5):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center" if col_idx in [2,3] else "left", vertical="center")
                
                for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4), start=2):
                    ws.row_dimensions[r_idx].height = 20
                    is_zebra = (r_idx % 2 == 0)
                    
                    for c_idx, cell in enumerate(row_cells, start=1):
                        cell.font = body_font
                        cell.border = thin_border
                        if is_zebra:
                            cell.fill = zebra_fill
                        
                        if c_idx == 1 or c_idx == 4:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif c_idx == 2:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0'
                        elif c_idx == 3:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        return True
    except Exception as e:
        print(f"  [ERROR] Failed to save Excel file {output_excel_path}: {e}")
        return False

def process_all_batch_reports():
    input_dir = "input"
    output_dir = "output"
    
    if not os.path.exists(input_dir):
        os.makedirs(input_dir, exist_ok=True)
        return
    os.makedirs(output_dir, exist_ok=True)
    
    files = [f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')]
    if not files:
        print("No PDF files found in 'input' directory.")
        return
        
    print(f"Found {len(files)} file(s) to process.\n")
    success_count = 0
    for filename in files:
        pdf_path = os.path.join(input_dir, filename)
        output_path = os.path.join(output_dir, f"{os.path.splitext(filename)[0]}.xlsx")
        
        print(f"-> Processing: {filename}")
        if process_single_pdf(pdf_path, output_path):
            print(f"   [SUCCESS] Saved to: {output_path}\n")
            success_count += 1
            
    print(f"Processing complete! Successfully transformed {success_count}/{len(files)} files.")

if __name__ == "__main__":
    process_all_batch_reports()